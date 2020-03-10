from openpyxl import load_workbook


def getRowCount(file):
    wb = load_workbook(file)
    sheet = wb.active
    return sheet.max_row

def getColumnCount(file):
    wb = load_workbook(file)
    sheet = wb.active
    return sheet.max_column

def getCellData(file, cell):
    wb = load_workbook(file)
    sheet = wb.active
    return sheet.cell(row=cell[0], column=cell[1]).value

def setCellData(file, cell, data):
    wb = load_workbook(file)
    sheet = wb.active
    sheet.cell(row=cell[0], column=cell[1]).value = data
    wb.save(file)


